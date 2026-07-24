import openpyxl, os, json, re

SHEET_CONFIGS = {
    'Main menu': {'name_col': 'I', 'desc_col': 'J', 'cc_col': 'K'},
    'Base panel allowances': {'name_col': 'K', 'desc_col': 'L', 'cc_col': 'M', 'formula_col': 'N'},
    'Lid panel allowances': {'name_col': 'K', 'desc_col': 'L', 'cc_col': 'M', 'formula_col': 'N'},
    'Allowances': {'name_col': 'B', 'desc_col': 'C', 'cc_col': 'D', 'formula_col': 'E'},
    'Base minor flap slot details': {'name_col': 'I', 'desc_col': 'J', 'formula_col': 'L'},
    'Base major flap slot details': {'name_col': 'I', 'desc_col': 'J', 'formula_col': 'L'},
    'Lid minor flap slot details ': {'name_col': 'I', 'desc_col': 'J', 'formula_col': 'L'},
    'Lid major flap slot details': {'name_col': 'I', 'desc_col': 'J', 'formula_col': 'L'},
    'Hidden menu1': {'name_col': 'B', 'desc_col': 'C', 'formula_col': 'E'},
    'KDF dimensions': {'name_col': 'B', 'desc_col': 'C', 'formula_col': 'E'},
}

def col_to_idx(col_letter):
    return openpyxl.utils.column_index_from_string(col_letter) - 1

def is_valid_var_name(s):
    if not s or len(s) > 20:
        return False
    s = s.strip()
    bad = ['Variable name','Description','C.C. / E.C.','Remarks','Means','Means "invisible variable"','C.C. = Check condition','E.C. = Enforced condition']
    if s in bad:
        return False
    return bool(re.match(r'^[A-Z][A-Z0-9_]*[A-Z0-9]$|^[A-Z]$', s))

def is_meaningful_formula(s):
    if not s:
        return False
    s = s.strip()
    bad = ['Remarks','Formula 1','Formula 2','Formula 3','Formula 4','Formula 5','Formula 6','Formula1','Formula2','Formula','Ref Formula 1','Ref Formula 2','Ref Formula 3','Ref Formula 4','Means "invisible variable"']
    if s in bad:
        return False
    if not re.search(r'[A-Z]', s):
        return False
    return True

docs_dir = 'C:/Users/Administrator/WorkBuddy/2026-07-23-13-09-20/FEFCO/Documentation'
files = sorted([f for f in os.listdir(docs_dir) if f.endswith('.xlsx')])

all_data = {}

for fname in files:
    box_id = fname.replace('Documentation ', '').replace('.xlsx','')
    try:
        wb = openpyxl.load_workbook(os.path.join(docs_dir, fname), data_only=True)
        info = {'sheets': wb.sheetnames, 'inputs': [], 'derived': [], 'formulas': {}}

        if 'Main menu' in wb.sheetnames:
            ws = wb['Main menu']
            cfg = SHEET_CONFIGS['Main menu']
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
                if len(row) <= max(col_to_idx(cfg['name_col']), col_to_idx(cfg['desc_col'])):
                    continue
                name = str(row[col_to_idx(cfg['name_col'])]).strip() if row[col_to_idx(cfg['name_col'])] else ''
                desc = str(row[col_to_idx(cfg['desc_col'])]).strip() if row[col_to_idx(cfg['desc_col'])] else ''
                cc = str(row[col_to_idx(cfg['cc_col'])]).strip() if row[col_to_idx(cfg['cc_col'])] else ''
                if is_valid_var_name(name) and desc and 'Description' not in desc:
                    info['inputs'].append({'var': name, 'desc': desc, 'cc': cc})

        for sheet_name, cfg in SHEET_CONFIGS.items():
            if sheet_name == 'Main menu' or sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            required_cols = [col_to_idx(c) for c in cfg.values() if c]
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
                if not row or len(row) <= max(required_cols):
                    continue
                name = str(row[col_to_idx(cfg['name_col'])]).strip() if row[col_to_idx(cfg['name_col'])] else ''
                desc = str(row[col_to_idx(cfg['desc_col'])]).strip() if cfg.get('desc_col') and row[col_to_idx(cfg['desc_col'])] else ''
                formula = str(row[col_to_idx(cfg['formula_col'])]).strip() if cfg.get('formula_col') and row[col_to_idx(cfg['formula_col'])] else ''
                cc = str(row[col_to_idx(cfg['cc_col'])]).strip() if cfg.get('cc_col') and row[col_to_idx(cfg['cc_col'])] else ''

                if is_valid_var_name(name):
                    if is_meaningful_formula(formula) or formula in ['CAL','IL','0']:
                        info['derived'].append({'var': name, 'desc': desc, 'formula': formula, 'sheet': sheet_name, 'cc': cc})
                        info['formulas'][name] = formula

        seen = set()
        unique_derived = []
        for d in info['derived']:
            if d['var'] not in seen:
                seen.add(d['var'])
                unique_derived.append(d)
        info['derived'] = unique_derived

        all_data[box_id] = info
    except Exception as e:
        all_data[box_id] = {'error': str(e)}

out_path = 'C:/Users/Administrator/WorkBuddy/2026-07-23-13-09-20/fefco_data_clean.json'
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump(all_data, f, indent=2, ensure_ascii=False)

print(f'Extracted clean data for {len(all_data)} box types')
print(f'Saved to {out_path}')

key_boxes = ['FT_0200','FT_0201','FT_0203','FT_0301','FT_0400','FT_0421','FT_0425','FT_0601','FT_0713','FT_0900','FT_0911']
for bid in key_boxes:
    if bid in all_data:
        info = all_data[bid]
        print(f'\n=== {bid} ===')
        print(f'  Inputs: {[d["var"] for d in info["inputs"]]}')
        print(f'  Derived: {[d["var"] for d in info["derived"]]}')
